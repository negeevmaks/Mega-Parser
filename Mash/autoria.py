import asyncio
from itertools import cycle
from bs4 import BeautifulSoup
import ui
from playwright.async_api import async_playwright
import openpyxl
from Mash.cars import Cars

base_url = 'https://auto.ria.com/uk/car/'

def check_model_exists(result_file, url, ws):
    for row in ws.iter_rows(min_row=2):
        cell_value = row[0].value

        if url in cell_value:
            return True

    return False

async def process_link(link, result_file, wb, ws, date, local):
    ws[f'A{1}'] = f'URL'
    ws[f'B{1}'] = f'Model'
    ws[f'C{1}'] = f'Price'
    ws[f'D{1}'] = f'Number'
    ws[f'E{1}'] = f'Locate'
    ws[f'F{1}'] = f'Seller'

    wb.save(f'{result_file}/autoria.xlsx')

    if check_model_exists(result_file, link, ws):
        print(link)
        print('skiped')
        return None

    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()

            context = await browser.new_context()
            page = await context.new_page()
            await page.goto(link)

            content = await page.content()
            soup = BeautifulSoup(content, 'html.parser')

            name = soup.find('h3', {'class': 'auto-content_title'})
            name = name.text if name else None

            locate = soup.find('div', {'class': 'item_inner'})
            locate = locate.text if locate else None

            seller = soup.find('div', {'class': 'seller_info_name'})
            seller = seller.text if seller else None

            data = int(name[-4:])

            if int(date[0]) <= int(data) <= int(date[1]):
                pass

            else:
                return 0

            name = name[:-4]

            price = soup.find('strong')

            if price:
                price_text = price.text

                if '$' in price_text:
                    price = price_text
                elif 'грн' in price_text:
                    price = price_text
                else:
                    return None
            else:
                price = None

            print(name, price, locate)

            # Extract phone number
            await page.click('.phone_show_link')
            await asyncio.sleep(1)
            number_element = await page.query_selector('.popup-successful-call-desk')
            number = await number_element.text_content() if number_element else None

            print(number)

            last_row = ws.max_row + 1

            ws[f'A{last_row}'] = link
            ws[f'B{last_row}'] = name
            ws[f'C{last_row}'] = price
            ws[f'D{last_row}'] = number
            ws[f'E{last_row}'] = locate
            ws[f'F{last_row}'] = seller
            ws[f'G{last_row}'] = data

            wb.save(f'{result_file}/autoria.xlsx')

            await browser.close()

        local += 1
        return local

    except Exception as e:
        print(f"Error processing link {link}: {str(e)}")
        return 1

async def main(Ca, value, result_file, category, data):
    wb = openpyxl.load_workbook(f'{result_file}/autoria.xlsx')
    ws = wb.active

    num = 0
    links = await Ca.base_requ(value, category)

    sub_string = "check_selection"

    linka = []

    for url in links:
        if sub_string not in url:
            linka.append(url)

    links = list(set(linka))
    print(links)

    cars_res = len(links)
    links = list(links)  # Convert set to a list

    local = 1

    for i in range(0, len(links)):
        link = links[i]
        tasks = [process_link(link, result_file, wb, ws, data, local)]
        local = await asyncio.gather(*tasks)
        print(local)

        num += 1
        await write_link(cars_res, num)

async def write_link(cars_res, num):
    print(f'Writing process: {num}/{cars_res}')
    ui.ui_update(cars_res, num)

def start_parsing(value, result_file, category, data):
    Ca = Cars()

    def create_workbook(path):
        workbook = openpyxl.Workbook()
        workbook.save(path)

    create_workbook(f'{result_file}/autoria.xlsx')

    loop = asyncio.get_event_loop()
    loop.run_until_complete(main(Ca, value, result_file, category, data))
