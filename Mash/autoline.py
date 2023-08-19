import requests
import ui
from bs4 import BeautifulSoup
import openpyxl
import os

base_url = 'https://autoline.ua/'

def get_category():
    response = requests.get(base_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    category = []

    ads = soup.find_all('a', {'class': 'item'})

    for ad in ads:
        if 'https://autoline.ua' in ad.get('href'):
            category.append(ad.get('href').replace('https://autoline.ua/-/', ''))

    return category

def pages(link):
    resuk = []
    resuk.append(f'https://autoline.ua/-/{link}')
    return set(list(resuk))

def get_links(category, cares, num_f=0):
    returned = []
    reh = []
    itemed = []

    for cat in category:
        link = set(list(pages(cat)))

        for li in link:
            response = requests.get(li)
            soup = BeautifulSoup(response.text, 'html.parser')
            result = soup.find_all('div', {'class': 'subitem'})

            for res in result:
                rem = res.find('a')
                reh.append(rem.get('href'))

            for i in range(0, num_f):
                reh.pop(0)

            for res in reh:
                response = requests.get(res)
                soup = BeautifulSoup(response.text, 'html.parser')
                mogaz = soup.find_all('div', {'class': 'item sales-list-item with-widget-price-extension with-widget-photo-widget item-active ecommerce-item'})

                for item in mogaz:
                    rek = item.find('a')
                    reg = rek.get('href')

                    itemed.append(reg)
                    print(reg)
                    print(len(itemed))

                    if len(itemed) >= cares:
                        return itemed

    return returned

def check_model_exists(url, ws):
    for row in ws.iter_rows(min_row=2):
        cell_value = row[0].value

        if url in cell_value:
            return True

    return False

def start_parsing(category, pric, cares, result_folder):
    
    def create_workbook(path):
        if not os.path.isfile(path):
            workbook = openpyxl.Workbook()
            workbook.save(path)

    workbook_path = f'{result_folder}/autoline.xlsx'
    create_workbook(workbook_path)

    try:
        wb = openpyxl.load_workbook(workbook_path)
    except openpyxl.utils.exceptions.InvalidFileException:
        os.remove(workbook_path)
        create_workbook(workbook_path)
        wb = openpyxl.load_workbook(workbook_path)
        
    ws = wb.active
    links = get_links(category, cares)

    end = []
    ens = 0
    last_row = 2
    num_f = 1

    for link in links:
        response = requests.get(link)
        soup = BeautifulSoup(response.text, 'html.parser')
        price = soup.find('div', {'class': 'value'}).text.replace('   ', '').replace('\n', '').replace('  ', '')

        if price != 'за запитом':
            try:
                if 'грн' in price:
                    if 'грн/зм' in price:
                        pride = int(price[:-9].replace(" ", "").replace("\xa0", "").replace(',', ''))
                    else:
                        pride = int(price[:-3].replace(" ", "").replace("\xa0", "").replace(',', ''))
                elif '—(аукціон' in price:
                    pride = int(price[:-9].replace(" ", "").replace("\xa0", "").replace(',', ''))
                else:
                    pride = int(price[:-1].replace(" ", "").replace("\xa0", "").replace(',', ''))
            except:
                pride = 0

            if pric[0] < pride and pride < pric[1]:
                if soup.find('span', {'class': 'loc-country'}).text == 'Україна':
                    if not check_model_exists(link, ws):
                        ws[f'A{last_row}'] = f'URL: {link}'
                        ws[f'B{last_row}'] = f"Model: {(soup.find('h1', {'class': 'sf-title'}).text)}"
                        ws[f'C{last_row}'] = f'Price: {price}'
                        ws[f'D{last_row}'] = f"Number: {' '.join(text.text.replace('        ', '') for text in soup.find_all('a', {'class': 'full-number'}))}"

                        wb.save(f'{result_folder}/autoline.xlsx')
                        last_row += 1
                        ens += 1
                        ui.ui_update(cares, ens)

        else:
            if soup.find('span', {'class': 'loc-country'}).text == 'Україна':
                if not check_model_exists(link, ws):
                    ws[f'A{last_row}'] = f'URL: {link}'
                    ws[f'B{last_row}'] = f"Model: {(soup.find('h1', {'class': 'sf-title'}).text)}"
                    ws[f'C{last_row}'] = f'Price: {price}'
                    ws[f'D{last_row}'] = f"Number: {' '.join(text.text.replace('        ', '') for text in soup.find_all('a', {'class': 'full-number'}))}"

                    wb.save(f'{result_folder}/autoline.xlsx')
                    last_row += 1
                    ens += 1
                    ui.ui_update(cares, ens)

    while True:
        num_f += 1
        if ens < cares:
            try:
                links = get_links([*(category + f'?page={num_f}' for category in category)], cares - ens - 1, num_f)
                for link in links:
                    response = requests.get(link)
                    soup = BeautifulSoup(response.text, 'html.parser')
                    price = soup.find('div', {'class': 'value'}).text.replace('   ', '').replace('\n', '').replace('  ', '')

                    if price != 'за запитом':
                        try:
                            if 'грн' in price:
                                if 'грн/зм' in price:
                                    pride = int(price[:-9].replace(" ", "").replace("\xa0", "").replace(',', ''))
                                else:
                                    pride = int(price[:-3].replace(" ", "").replace("\xa0", "").replace(',', ''))
                            elif '—(аукціон' in price:
                                pride = int(price[:-9].replace(" ", "").replace("\xa0", "").replace(',', ''))
                            else:
                                pride = int(price[:-1].replace(" ", "").replace("\xa0", "").replace(',', ''))
                        except:
                            pride = 0

                        if pric[0] < pride and pride < pric[1]:
                            if soup.find('span', {'class': 'loc-country'}).text == 'Україна':
                                if not check_model_exists(link, ws):
                                    ws[f'A{last_row}'] = f'URL: {link}'
                                    ws[f'B{last_row}'] = f"Model: {(soup.find('h1', {'class': 'sf-title'}).text)}"
                                    ws[f'C{last_row}'] = f'Price: {price}'
                                    ws[f'D{last_row}'] = f"Number: {' '.join(text.text.replace('        ', '') for text in soup.find_all('a', {'class': 'full-number'}))}"

                                    wb.save(f'{result_folder}/autoline.xlsx')
                                    last_row += 1
                                    ens += 1
                                    ui.ui_update(cares, ens)
                    else:
                        if soup.find('span', {'class': 'loc-country'}).text == 'Україна':
                            if not check_model_exists(link, ws):
                                ws[f'A{last_row}'] = f'URL: {link}'
                                ws[f'B{last_row}'] = f"Model: {(soup.find('h1', {'class': 'sf-title'}).text)}"
                                ws[f'C{last_row}'] = f'Price: {price}'
                                ws[f'D{last_row}'] = f"Number: {' '.join(text.text.replace('        ', '') for text in soup.find_all('a', {'class': 'full-number'}))}"

                                wb.save(f'{result_folder}/autoline.xlsx')
                                last_row += 1
                                ens += 1
                                ui.ui_update(cares, ens)
            except:
                break
        else:
            break

    ui.ui_update(1, 1)
